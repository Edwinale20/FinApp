import streamlit as st
import pandas as pd
from pathlib import Path
from io import BytesIO
import requests
import io
import openpyxl


st.title("🤖 Resumen semanal")
st.markdown("La finalidad de esta segunda parte, es ingresar todo tip de movimiento registrado en mi día a día", unsafe_allow_html=True)

# -----------------------------------------------------------------------------------------------------------------------------

cfg = st.secrets["onedrive"]
CLIENT_ID = cfg["client_id"]
CLIENT_SECRET = cfg["client_secret"]
REFRESH_TOKEN = cfg["refresh_token"]
REDIRECT_URI = cfg["redirect_uri"]

def get_access_token():
    url = "https://login.microsoftonline.com/consumers/oauth2/v2.0/token"
    data = {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "refresh_token": REFRESH_TOKEN,
        "grant_type": "refresh_token",
        "redirect_uri": REDIRECT_URI,
        "scope": "Files.ReadWrite Files.Read.All User.Read offline_access"
    }
    r = requests.post(url, data=data)
    return r.json()

@st.cache_data
def list_excel_files(access_token):
    url = "https://graph.microsoft.com/v1.0/me/drive/root:/FinApp:/children"
    headers = {"Authorization": f"Bearer {access_token}"}
    r = requests.get(url, headers=headers).json()

    return [f for f in r.get("value", []) if f["name"].lower().endswith(".xlsx")]

@st.cache_data
def download_excel_df(access_token, file_id):
    url = f"https://graph.microsoft.com/v1.0/me/drive/items/{file_id}/content"
    headers = {"Authorization": f"Bearer {access_token}"}
    content = requests.get(url, headers=headers).content
    return pd.read_excel(io.BytesIO(content))


# -----------------------------------------------------------------------------------------------------------------------------

token = get_access_token()

if "access_token" not in token:
    st.error("❌ Error obteniendo access_token")
    st.code(token)
else:
    access_token = token["access_token"]

    files = list_excel_files(access_token)

    # Buscar el archivo específico llamado Tracking.xlsx
    tracking_file = next((f for f in files if f["name"] == "Tracking.xlsx"), None)
    
    if tracking_file:
        df_tracking = download_excel_df(access_token, tracking_file["id"])
        #df_tracking2 = download_excel_df(access_token, tracking_file["id"], sheet_name="Deudas")  # segunda hoja
        st.success("✅ Archivo 'Tracking.xlsx' cargado correctamente.")
    else:
        st.error("❌ No se encontró el archivo 'Tracking.xlsx' en la carpeta FinApp.")

# -----------------------------------------------------------------------------------------------------------------------------

def append_row_to_onedrive_excel(access_token, file_id, sheet_name, row_dict):
    url = f"https://graph.microsoft.com/v1.0/me/drive/items/{file_id}/content"

    # 1) Descargar xlsx
    r = requests.get(url, headers={"Authorization": f"Bearer {access_token}"})
    r.raise_for_status()
    xlsx_bytes = r.content

    # 2) Abrir + append 1 fila en la hoja
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb[sheet_name]

    headers = [c.value for c in ws[1]]  # fila 1 = headers
    col = {h: i+1 for i, h in enumerate(headers) if h}

    next_row = ws.max_row + 1
    for k, v in row_dict.items():
        ws.cell(row=next_row, column=col[k], value=v)

    # 3) Guardar a bytes
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    # 4) Subir de vuelta el MISMO archivo
    r2 = requests.put(
        url,
        headers={
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
        data=out.getvalue(),
    )
    r2.raise_for_status()

# Inputs
Nombre = st.text_input("🖋️ Ingresa la Descripción:")
Cantidad = st.number_input("💲Ingresa el monto:", min_value=0.0, step=1.0)
Categoria = st.text_input("🍻 Ingresa la categoría:")
fecha = st.date_input("🗓️ Selecciona la fecha:")

Submit = st.button("Ingresar")

if Submit and tracking_file:
    new_row = {
        "Fecha": fecha,                 # st.date_input ya da date
        "Categoría": Categoria,
        "Descripción": Nombre,
        "Monto": float(Cantidad),
        # "Concepto": "Gasto",          # si tu hoja lo tiene, agrégalo
    }

    append_row_to_onedrive_excel(
        access_token,
        tracking_file["id"],
        "Registro",
        new_row
    )

    st.cache_data.clear()   # IMPORTANTÍSIMO: si no, verás el excel viejo
    st.success("✅ Se agregó la nueva línea en OneDrive (Tracking.xlsx → Registro)")
    st.rerun()

